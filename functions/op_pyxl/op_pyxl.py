import pandas as pd
import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils.dataframe import dataframe_to_rows

class op_pyxl:
    def __init__(self, path):
        # data_only= True (i.e. values i/o formulas)
        self.readwb = openpyxl.load_workbook(filename= path, read_only= False, data_only= True,
                                           keep_vba= True, keep_links= True)
        # turn wb to write-friendly type: read_only= False and data_only= False (i.e. save Excel file with formulas)
        self.writewb = openpyxl.load_workbook(filename=path, read_only=False, data_only=False,
                                         keep_vba=True, keep_links=True)
        self.wb = self.readwb
        self.path = path

    def get_tbl_as_df(self, tbl_name):
        # step 1: define list names.
        values = []
        tblvallist = []
        # step 2: ws > tbl > each row in tbl > each cell value in row
        for ws in self.wb.worksheets:
            for tbl in ws._tables:
                if tbl.displayName == tbl_name:
                    cellrefs = ws[tbl.ref]
                    for cellrows in cellrefs:
                        for cell in cellrows:
                            values.append(cell.value)
                        tblvallist.append(list(values))
                        values.clear()
        return pd.DataFrame(tblvallist)

    def rng(self, rngname):
        rngdata = list(self.wb.defined_names[rngname].destinations)[0]  # sheet name, cell address
        sheet = rngdata[0]  # sheet name
        address = rngdata[1].replace('$', '')  # address
        return self.wb[sheet][address]

    def update_rng_val(self, xlitem, newval):
        # turn the master wb into write-mode, then turn it back to read-mode for other functions.
        self.wb = self.writewb
        if isinstance(newval,pd.DataFrame):
            for r in dataframe_to_rows(newval, index=True, header=True):
                self.wb[xlitem].append(r)
        else:
            self.rng(xlitem).value = newval
        self.writewb.save(self.path)
        self.wb = self.readwb

    # back-up code
    def update_rng_val_old(self, rngname, newval):
        # turn the master wb into write-mode, then turn it back to read-mode for other functions.
        self.wb = self.writewb
        self.rng(rngname).value = newval
        self.writewb.save(self.path)
        self.wb = self.readwb

    # in list myrng, find the lookup val in lookup rng and return a corressponding value in the result column.
    def vlookup(self, myrng, lookup_rng, lookup_val, result_rng):
        i = 0
        for header in myrng.iloc[i]:
            if header == lookup_rng:
                lookupcol = i
            elif header == result_rng:
                resultcol = i
            i += 1
        i = 0
        for item in myrng[lookupcol]:
            if item == lookup_val:
                resultrow = i
            i += 1
        return myrng[resultcol][resultrow]

